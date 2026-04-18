from __future__ import annotations

import calendar
from pathlib import Path

import logging

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from schedule_askue.core.calendar_ua import UkrainianCalendar
from schedule_askue.core.project_config import load_project_config
from schedule_askue.core.shift_codes import normalize_shift_code
from schedule_askue.db.models import Employee

logger = logging.getLogger(__name__)


class ExcelExporter:
    MONTH_NAMES_UA = {
        1: "СІЧЕНЬ",
        2: "ЛЮТИЙ",
        3: "БЕРЕЗЕНЬ",
        4: "КВІТЕНЬ",
        5: "ТРАВЕНЬ",
        6: "ЧЕРВЕНЬ",
        7: "ЛИПЕНЬ",
        8: "СЕРПЕНЬ",
        9: "ВЕРЕСЕНЬ",
        10: "ЖОВТЕНЬ",
        11: "ЛИСТОПАД",
        12: "ГРУДЕНЬ",
    }
    WEEKDAY_LABELS = {0: "пн", 1: "вт", 2: "ср", 3: "чт", 4: "пт", 5: "сб", 6: "нд"}

    def __init__(self, project_root: Path) -> None:
        self.project_root = project_root
        self.config = load_project_config(self.project_root)
        self.shift_colors = {
            code: meta.get("color", "").replace("#", "")
            for code, meta in self.config.get("shifts", {})
            .get("definitions", {})
            .items()
        }
        thin = Side(style="thin", color="000000")
        self.thin_side = thin
        self.thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        self.font_regular = Font(name="Arial Cyr", size=12)
        self.font_bold = Font(name="Arial Cyr", size=12, bold=True)
        self.font_title = Font(name="Arial Cyr", size=14, bold=True)
        self.font_note = Font(name="Arial Cyr", size=10)
        self.font_weekend = Font(name="Arial Cyr", size=12, color="C91C23", bold=True)
        self.weekend_fill = PatternFill("solid", fgColor="FFCCCC")

    def export_month(
        self,
        output_path: Path,
        year: int,
        month: int,
        employees: list[Employee],
        assignments: dict[int, dict[int, str]],
        settings: dict[str, str],
        calendar_ua: UkrainianCalendar,
    ) -> Path:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = f"{month:02d}-{year}"

        month_days = calendar.monthrange(year, month)[1]
        month_info = calendar_ua.get_month_info(year, month)

        self._setup_dimensions(sheet, month_days)
        self._build_header(sheet, year, month, settings, month_days)
        self._build_schedule_table(sheet, employees, assignments, month_info)
        self._build_note_and_signatures(sheet, employees, month_days)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        return output_path

    def _setup_dimensions(self, sheet, month_days: int) -> None:
        last_day_col = month_days + 1
        trailing_col = last_day_col + 3

        sheet.column_dimensions["A"].width = 19.66
        for index in range(2, trailing_col + 1):
            sheet.column_dimensions[get_column_letter(index)].width = 4.57

        row_heights = {
            1: 15.6,
            2: 15.6,
            3: 15.6,
            4: 15.6,
            5: 15.0,
            8: 17.4,
            9: 15.6,
            11: 15.0,
            12: 15.0,
            13: 15.75,
            14: 15.75,
            15: 15.75,
            16: 15.75,
            19: 3.0,
            20: 15.75,
            21: 15.75,
            22: 15.75,
            23: 15.6,
            24: 15.6,
            25: 15.6,
            26: 15.0,
            27: 14.25,
            28: 15.0,
            29: 15.0,
            30: 15.0,
            31: 13.5,
            32: 15.0,
            33: 15.0,
        }
        for row, height in row_heights.items():
            sheet.row_dimensions[row].height = height

    def _build_header(
        self, sheet, year: int, month: int, settings: dict[str, str], month_days: int
    ) -> None:
        last_day_col = month_days + 1
        last_day_letter = get_column_letter(last_day_col)
        trailing_letter = get_column_letter(last_day_col + 3)

        sheet.merge_cells(f"S6:{trailing_letter}6")
        sheet["Y1"] = '" ЗАТВЕРДЖУЮ"'
        sheet["Y2"] = settings.get("director_title", "Заступник директора")
        sheet["Y3"] = settings.get("company_name", "")
        sheet["Y4"] = f"____________{settings.get('director_name', '')}"
        for ref in ["Y1", "Y2", "Y3", "Y4"]:
            sheet[ref].font = self.font_regular

        sheet.merge_cells(f"A8:{last_day_letter}8")
        sheet["A8"] = settings.get("schedule_title", "Графік роботи інженерів АСКУЕ")
        sheet["A8"].font = self.font_title
        sheet["A8"].alignment = Alignment(horizontal="center")

        sheet.merge_cells(f"A9:{last_day_letter}9")
        sheet["A9"] = f"{self.MONTH_NAMES_UA[month]} {year} р."
        sheet["A9"].font = self.font_bold
        sheet["A9"].alignment = Alignment(horizontal="center")

    def _build_schedule_table(
        self,
        sheet,
        employees: list[Employee],
        assignments: dict[int, dict[int, str]],
        month_info: dict[int, object],
    ) -> None:
        month_days = len(month_info)

        sheet.merge_cells("A11:A12")
        anchor = sheet["A11"]
        anchor.value = "ПІБ співробітника"
        anchor.font = self.font_regular
        anchor.alignment = Alignment(horizontal="center", vertical="center")
        anchor.border = self.thin_border

        for day in range(1, month_days + 1):
            col = day + 1
            top_cell = sheet.cell(row=11, column=col, value=f"{day:02d}")
            bottom_cell = sheet.cell(
                row=12, column=col, value=self.WEEKDAY_LABELS[month_info[day].weekday]
            )

            is_weekend = month_info[day].is_weekend

            for cell in (top_cell, bottom_cell):
                cell.font = self.font_weekend if is_weekend else self.font_regular
                cell.alignment = Alignment(horizontal="center", vertical="center")
            top_cell.border = Border(
                left=self.thin_side, right=self.thin_side, top=self.thin_side
            )
            bottom_cell.border = Border(
                left=self.thin_side, right=self.thin_side, bottom=self.thin_side
            )
            if is_weekend:
                top_cell.fill = self.weekend_fill
                bottom_cell.fill = self.weekend_fill
            else:
                # Звичайні дні - білий фон
                white_fill = PatternFill("solid", fgColor="FFFFFF")
                top_cell.fill = white_fill
                bottom_cell.fill = white_fill

        start_row = 13
        for row_offset, employee in enumerate(employees):
            row = start_row + row_offset
            name_cell = sheet.cell(row=row, column=1, value=employee.short_name)
            name_cell.font = self.font_regular
            name_cell.alignment = Alignment(horizontal="left", vertical="center")
            name_cell.border = self.thin_border

            employee_days = assignments.get(employee.id or 0, {})
            for day in range(1, month_days + 1):
                shift_code = normalize_shift_code(employee_days.get(day, ""))
                cell = sheet.cell(
                    row=row, column=day + 1, value=self._display_shift(shift_code)
                )

                # Визначити, чи день вихідний/святковий
                is_special_day = month_info[day].is_weekend

                # Шрифт: жирний для "Д", червоний текст для вихідних днів
                if shift_code == "Д":
                    if is_special_day:
                        cell.font = Font(
                            name="Arial Cyr", size=12, bold=True, color="C91C23"
                        )
                    else:
                        cell.font = self.font_bold
                else:
                    if is_special_day:
                        cell.font = Font(name="Arial Cyr", size=12, color="C91C23")
                    else:
                        cell.font = self.font_regular

                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = self.thin_border

                # Фон залежить від коду зміни
                fill_color = self.shift_colors.get(shift_code)
                if fill_color:
                    cell.fill = PatternFill("solid", fgColor=fill_color)

    def _build_note_and_signatures(
        self, sheet, employees: list[Employee], month_days: int
    ) -> None:
        start_row = 13 + len(employees)
        current_row = start_row

        # Примітка про чергування
        duty_note = self.config.get("export", {}).get(
            "duty_note",
            "Примітка: Д — Чергування, робота з перервою, вранці 5 год (08:00–13:00), вночі 3 год (22:00–01:00)",
        )
        last_day_letter = get_column_letter(month_days + 1)
        sheet.merge_cells(f"B{current_row}:{last_day_letter}{current_row}")
        note_cell = sheet.cell(row=current_row, column=2, value=duty_note)
        note_cell.font = Font(name="Arial Cyr", size=10)
        note_cell.alignment = Alignment(horizontal="left", vertical="center")
        sheet.row_dimensions[current_row].height = 13.5
        current_row += 1

        # Порожній рядок
        current_row += 1

        # Умовні позначення
        sheet.cell(row=current_row, column=2, value="Умовні позначення:").font = Font(
            name="Arial Cyr", size=10
        )
        sheet.row_dimensions[current_row].height = 13.5
        current_row += 1

        # Отримати умовні позначення з конфігу
        legend = self.config.get("export", {}).get(
            "legend",
            [
                {"code": "Д", "description": "— Чергування"},
                {"code": "Р", "description": "— Робочий день"},
                {"code": "В", "description": "— Вихідний"},
                {"code": "О", "description": "— Відпустка"},
            ],
        )

        for item in legend:
            code = item.get("code", "")
            description = item.get("description", "")

            # Стовпець B: код з кольоровим фоном
            code_cell = sheet.cell(row=current_row, column=2, value=code)
            code_cell.font = Font(name="Arial Cyr", size=10, bold=True)
            code_cell.alignment = Alignment(horizontal="center", vertical="center")
            fill_color = self.shift_colors.get(code)
            if fill_color:
                code_cell.fill = PatternFill("solid", fgColor=fill_color)

            # Стовпець C: опис
            desc_cell = sheet.cell(row=current_row, column=3, value=description)
            desc_cell.font = Font(name="Arial Cyr", size=10)
            desc_cell.alignment = Alignment(horizontal="left", vertical="center")

            sheet.row_dimensions[current_row].height = 13.5
            current_row += 1

        # Порожній рядок
        current_row += 1

        # Ознайомлені
        sheet.cell(row=current_row, column=2, value="Ознайомлені:").font = Font(
            name="Arial Cyr", size=10
        )
        sheet.row_dimensions[current_row].height = 13.2
        current_row += 1

        # Список співробітників з лініями для підпису
        for employee in employees:
            name_cell = sheet.cell(row=current_row, column=2, value=employee.short_name)
            name_cell.font = Font(name="Arial Cyr", size=10)
            name_cell.alignment = Alignment(horizontal="left", vertical="center")

            # Лінія для підпису в стовпці E - вирівняна по лівому краю
            signature_cell = sheet.cell(
                row=current_row, column=5, value="_______________________"
            )
            signature_cell.font = Font(name="Arial Cyr", size=10)
            signature_cell.alignment = Alignment(horizontal="left", vertical="center")

            sheet.row_dimensions[current_row].height = 13.2
            current_row += 1

    def _display_shift(self, shift_code: str) -> str:
        mapping = {"Р": "Р", "Д": "Д", "В": "В", "О": "О", "": ""}
        return mapping.get(shift_code, shift_code)
